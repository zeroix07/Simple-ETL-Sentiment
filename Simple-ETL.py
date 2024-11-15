#Simple ETL pipeline to analysis Sentiment response time and word count from PostgreSQL database and save to Excel file.
import psycopg2
import pandas as pd
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment

def get_database_connection():
    """Create and return a database connection."""
    try:
        conn = psycopg2.connect(
            dbname="your_dbname",
            user="your_username",
            password="your_password",
            host="ypur_host",
            port="your_port"
        )
        return conn
    except Exception as e:
        print(f"Error connecting to database: {str(e)}")
        raise

def extract_to_csv():
    """Extract data from PostgreSQL and save to CSV."""
    conn = None
    try:
        conn = get_database_connection()
        
        query = """
        SELECT s.* 
        FROM sentiment s
        """
        
        df = pd.read_sql_query(query, conn)
        
        # Create extract folder if doesn't exist
        os.makedirs('extract', exist_ok=True)
        
        # Save to CSV with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_file = f'extract/sentiment_{timestamp}.csv'
        df.to_csv(csv_file, index=False)
        
        print(f"Data extracted to {csv_file}")
        return csv_file
        
    except Exception as e:
        print(f"Error during extraction: {str(e)}")
        raise
    finally:
        if conn:
            conn.close()

def transform_data(csv_file):
    """Transform sentiment data for analysis."""
    try:
        df = pd.read_csv(csv_file)
        
        # Convert dates to datetime
        date_columns = ['at', 'repliedat']
        for col in date_columns:
            df[col] = pd.to_datetime(df[col])
        
        # Calculate response time and word count
        df['response_time_minutes'] = (df['repliedat'] - df['at']).dt.total_seconds() / 60
        df['word_count'] = df['text_final'].str.split().str.len()
        
        # Map sentiment categories
        df['sentiment_category'] = df['polarity'].map({'positive': 'Positive', 'negative': 'Negative'})
        
        # Calculate metrics
        metrics = {
            'total_reviews': len(df),
            'positive_reviews': (df['polarity'] == 'positive').sum(),
            'negative_reviews': (df['polarity'] == 'negative').sum(),
            'avg_response_time': df['response_time_minutes'].mean(),
            'avg_word_count': df['word_count'].mean(),
            'response_rate': (df['replycontent'].notna().sum() / len(df)) * 100
        }
        
        return df, metrics
        
    except Exception as e:
        print(f"Error during transformation: {str(e)}")
        raise

def load_to_excel(df, metrics, output_folder='analysis_output'):
    """Save analysis to Excel file."""
    try:
        os.makedirs(output_folder, exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f'{output_folder}/sentiment_analysis_{timestamp}.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sentiment Analysis'
        
        # Add headers with styling
        headers = ['Review ID', 'Content', 'Sentiment', 'Response Time (min)', 'Word Count']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for row_idx, row in enumerate(df.iterrows(), 2):
            row = row[1]
            ws.cell(row=row_idx, column=1, value=row['reviewid'])
            ws.cell(row=row_idx, column=2, value=row['content'])
            ws.cell(row=row_idx, column=3, value=row['sentiment_category'])
            ws.cell(row=row_idx, column=4, value=round(row['response_time_minutes'], 2))
            ws.cell(row=row_idx, column=5, value=row['word_count'])
        
        # Add metrics sheet
        ws_metrics = wb.create_sheet('Metrics')
        ws_metrics.append(['Metric', 'Value'])
        for metric, value in metrics.items():
            ws_metrics.append([metric, round(value, 2) if isinstance(value, float) else value])
        
        # Auto-adjust columns
        for worksheet in wb.worksheets:
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
        
        wb.save(excel_file)
        print(f"Analysis saved to {excel_file}")
        return True
        
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        raise

def run_etl():
    """Main ETL pipeline."""
    try:
        # Extract from DB to CSV
        csv_file = extract_to_csv()
        
        # Transform CSV data
        transformed_df, metrics = transform_data(csv_file)
        
        # Load to Excel
        load_to_excel(transformed_df, metrics)
        
        print("\nETL pipeline completed successfully")
        
    except Exception as e:
        print(f"ETL pipeline failed: {str(e)}")
        raise

if __name__ == "__main__":
    run_etl()